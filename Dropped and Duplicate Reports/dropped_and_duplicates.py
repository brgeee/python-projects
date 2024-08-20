import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

#load excel
file_path = 'C:/Users/EB1801445/OneDrive - Oracle Corporation/Documents/Documents/Useful Python Scripts/Dropped and Duplicate Reports/newcolumns.xlsx'
df = pd.read_excel(file_path)

#output columns
columns_to_output = ['Code System', 'Code', 'Names', 'Old Concept Name', 'New Concept Name']

#function - all duplicates (red, white, and green)
def process_all_dups(df):
    df['Normalized_Code'] = df['Code'].astype(str).str.rstrip('.0')
    grouped = df.groupby(['Code System', 'Normalized_Code'])
    matching_codes = []
    for (code_system, code), group in grouped:
        removed_from_concept = group['Old Concept Name'].notnull() & group['New Concept Name'].isnull()
        added_to_concept = group['Old Concept Name'].isnull() & group['New Concept Name'].notnull()
        remains_in_concept = group['Old Concept Name'].notnull() & group['New Concept Name'].notnull()
        if removed_from_concept.any() and added_to_concept.any() and remains_in_concept.any():
            matching_codes.append((code_system, code))
    matching_codes_df = df[(df[['Code System', 'Normalized_Code']].apply(tuple, axis=1).isin(matching_codes))]
    return matching_codes_df[columns_to_output]

#function - green/white duplicates
def process_gw_dups(df):
    removed_codes = df[df['Old Concept Name'].notnull() & df['New Concept Name'].isnull()]['Code'].unique()
    initial_matches = df[df['Old Concept Name'].notnull() & df['New Concept Name'].notnull()]
    duplicates = []
    for _, group in initial_matches.groupby(['Code System', 'Code']):
        initial_match_rows = group
        subsequent_rows = df[(df['Code System'] == group['Code System'].iloc[0]) & 
                             (df['Code'] == group['Code'].iloc[0]) & 
                             (df.index > initial_match_rows.index.max())]
        if (subsequent_rows['Old Concept Name'].isnull() & subsequent_rows['New Concept Name'].notnull()).any():
            duplicates.append((group['Code System'].iloc[0], group['Code'].iloc[0]))
    duplicates = [dup for dup in duplicates if dup[1] not in removed_codes]
    duplicate_codes_df = df[(df[['Code System', 'Code']].apply(tuple, axis=1).isin(duplicates))]
    return duplicate_codes_df[duplicate_codes_df['New Concept Name'].notnull()][columns_to_output]

#function - red/white duplicates
def process_rw_dups(df):
    df['Code'] = df['Code'].astype(str)
    added_codes = df[df['Old Concept Name'].isnull() & df['New Concept Name'].notnull()]['Code'].unique()
    df = df[~df['Code'].isin(added_codes)]
    initial_matches = df[df['Old Concept Name'].notnull() & df['New Concept Name'].notnull()]
    matching_codes = []
    for _, group in initial_matches.groupby(['Code System', 'Code']):
        subsequent_rows = df[(df['Code System'] == group['Code System'].iloc[0]) & (df['Code'] == group['Code'].iloc[0])]
        if (subsequent_rows['Old Concept Name'].notnull() & subsequent_rows['New Concept Name'].isnull()).any():
            matching_codes.append((group['Code System'].iloc[0], group['Code'].iloc[0]))
    matching_codes_df = df[df[['Code System', 'Code']].apply(tuple, axis=1).isin(matching_codes)]
    return matching_codes_df[columns_to_output]

#function - dropped
def process_dropped_codes(df):
    removed_not_added = []
    for code in df['Code'].unique():
        code_df = df[df['Code'] == code]
        if code_df['Old Concept Name'].notnull().any() and not code_df['New Concept Name'].notnull().any():
            removed_not_added.append(code)
    removed_not_added_df = df[df['Code'].isin(removed_not_added)]
    columns_to_output_dropped = ['Code System', 'Code', 'Names', 'Old Concept Name']
    return removed_not_added_df[columns_to_output_dropped]

#function - code additions
def process_code_additions(df):
    added_codes_df = df[df['Old Concept Name'].isnull() & df['New Concept Name'].notnull()]
    columns_to_output_additions = ['Code System', 'Code', 'Names', 'New Concept Name']
    return added_codes_df[columns_to_output_additions]

#process data
all_dups_df = process_all_dups(df)
gw_dups_df = process_gw_dups(df)
rw_dups_df = process_rw_dups(df)
dropped_codes_df = process_dropped_codes(df)
code_additions_df = process_code_additions(df)

#load workbook
wb = load_workbook(file_path)

#create dropped sheet
if "Dropped" in wb.sheetnames:
    wb.remove(wb["Dropped"])
ws_dropped = wb.create_sheet("Dropped", 1)

#dropped codes data and styles
for col, column_name in enumerate(dropped_codes_df.columns, start=1):
    ws_dropped.cell(row=1, column=col, value=column_name)
    ws_dropped.cell(row=1, column=col).font = Font(bold=True)
for row, data_row in enumerate(dropped_codes_df.itertuples(index=False), start=2):
    for col, value in enumerate(data_row, start=1):
        ws_dropped.cell(row=row, column=col, value=value)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color='9C0006')
for row in ws_dropped.iter_rows(min_row=2, max_row=ws_dropped.max_row, min_col=1, max_col=ws_dropped.max_column):
    old_concept = row[3].value
    if old_concept:
        for cell in row:
            cell.fill = red_fill
            cell.font = red_font

#create duplicates sheet
if "Duplicates" in wb.sheetnames:
    wb.remove(wb["Duplicates"])
ws = wb.create_sheet("Duplicates")

#write data for duplicates sheet w/headers
def write_data_with_header(ws, start_row, header, data):
    ws.cell(row=start_row, column=1, value=header)
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(columns_to_output))
    
    for col, column_name in enumerate(columns_to_output, start=1):
        ws.cell(row=start_row + 1, column=col, value=column_name)
        ws.cell(row=start_row + 1, column=col).font = Font(bold=True)
    
    for row, data_row in enumerate(data.itertuples(index=False), start=start_row + 2):
        for col, value in enumerate(data_row, start=1):
            ws.cell(row=row, column=col, value=value)
    
    return start_row + data.shape[0] + 3 

#duplicates sheet data and styles
next_row = write_data_with_header(ws, 1, "Green, White, and Red Duplicates", all_dups_df)
next_row = write_data_with_header(ws, next_row, "Green and White Duplicates", gw_dups_df)
next_row = write_data_with_header(ws, next_row, "Red and White Duplicates", rw_dups_df)

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color='006100')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    old_concept = row[3].value
    new_concept = row[4].value
    
    if isinstance(old_concept, str) and not isinstance(new_concept, str):
        for cell in row:
            cell.fill = red_fill
            cell.font = red_font
    elif not isinstance(old_concept, str) and isinstance(new_concept, str):
        for cell in row:
            cell.fill = green_fill
            cell.font = green_font

#create code additions sheet and styles
if "Additions" in wb.sheetnames:
    wb.remove(wb["Additions"])
ws_added = wb.create_sheet("Additions", 2)
for col, column_name in enumerate(code_additions_df.columns, start=1):
    ws_added.cell(row=1, column=col, value=column_name)
    ws_added.cell(row=1, column=col).font = Font(bold=True)
for row, data_row in enumerate(code_additions_df.itertuples(index=False), start=2):
    for col, value in enumerate(data_row, start=1):
        ws_added.cell(row=row, column=col, value=value)

for row in ws_added.iter_rows(min_row=2, max_row=ws_added.max_row, min_col=1, max_col=ws_added.max_column):
    for cell in row:
        cell.fill = green_fill
        cell.font = green_font

#save workbook
wb.save(file_path)
print(f"Data has been written to {file_path} in the 'Dropped', 'Duplicates', and 'Code Additions' sheets.")
